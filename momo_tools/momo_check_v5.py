# -*- coding: utf-8 -*-
"""
 m o m o _ c h e c k _ v 5
 在 v1/v2/v3/v4 基礎上改為 CSV 匯入、Excel 匯出（專抓手機版頁面、強化 timeout 重試）
"""

import datetime
import html
import re
import time
from typing import Dict, List
from urllib.parse import parse_qs, urlparse

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

MRDT_REGEX = re.compile(r"[MRDT][A-Za-z0-9]{5}", re.IGNORECASE)


def _find_bsmi_code(text: str) -> str:
    """
    從文字中找出第一組「合理的」商檢字號：
    - 基本格式：M/R/D/T + 5 碼英數字
    - 後 5 碼至少要含 1 個數字，避免誤抓純英文
    - 前一個字元若是 '#'（例如 #D62872 搜尋用 tag）則略過
    """
    if not text:
        return ""

    for m in MRDT_REGEX.finditer(text):
        start = m.start()
        code = m.group(0).upper()

        if start > 0 and text[start - 1] == "#":
            continue

        tail = code[1:]
        if not any(ch.isdigit() for ch in tail):
            continue

        return code

    return ""


def _extract_i_code(url: str) -> str:
    try:
        parsed = urlparse(url)
        qs = parse_qs(parsed.query)
        if "i_code" in qs and qs["i_code"]:
            return qs["i_code"][0]
    except Exception:
        pass
    return ""


def _parse_bsmi_from_soup(soup: BeautifulSoup) -> str:
    # 1) 含「認證字號」文字的區塊
    cert_block = soup.find(
        lambda tag: tag.name in ("th", "td", "div", "li") and "認證字號" in tag.get_text(strip=True)
    )
    if cert_block is not None:
        text = cert_block.get_text(" ", strip=True)
        code = _find_bsmi_code(text)
        if not code:
            sib = cert_block.find_next_sibling()
            if sib is not None:
                text2 = sib.get_text(" ", strip=True)
                code = _find_bsmi_code(text2)
        if code:
            return code

    # 2) 含「商檢字號」文字的區塊
    sj_block = soup.find(
        lambda tag: tag.name in ("th", "td", "li", "div", "span", "p") and "商檢字號" in tag.get_text(strip=True)
    )
    if sj_block is not None:
        text = sj_block.get_text(" ", strip=True)
        code = _find_bsmi_code(text)
        if not code:
            sib = sj_block.find_next_sibling()
            if sib is not None:
                text2 = sib.get_text(" ", strip=True)
                code = _find_bsmi_code(text2)
        if code:
            return code

    # 3) 舊的 panel-2 規格區塊邏輯保留
    spec_panel = soup.find("div", id="panel-2")
    if spec_panel is not None:
        label_div = spec_panel.find(lambda tag: tag.name == "div" and tag.get_text(strip=True) == "商檢字號")
        if label_div is not None:
            value_div = label_div.find_next_sibling("div")
            if value_div is not None:
                text = value_div.get_text(" ", strip=True)
                code = _find_bsmi_code(text)
                if code:
                    return code

        text = spec_panel.get_text(" ", strip=True)
        code = _find_bsmi_code(text)
        if code:
            return code

    # 4) 整頁全文 fallback
    full_text = soup.get_text(" ", strip=True)
    code = _find_bsmi_code(full_text)
    if code:
        return code

    return ""


# ➤ 對 MOMO 商品進行 retry + timeout 的簡易爬蟲（專抓手機版）
def parse_momo_simple(url: str, max_retries: int = 5) -> Dict[str, str]:
    headers = {"User-Agent": "Mozilla/5.0"}
    name = "未取得"
    prod_no = "未取得"
    zhigui_value = ""

    i_code = _extract_i_code(url)
    m_url = f"https://m.momoshop.com.tw/goods.momo?i_code={i_code}" if i_code else url

    for attempt in range(1, max_retries + 1):
        try:
            res = requests.get(m_url, headers=headers, timeout=20)
            res.raise_for_status()
            soup = BeautifulSoup(res.text, "html.parser")

            name_tag = soup.select_one("meta[property='og:title']")
            name = name_tag["content"].strip() if name_tag else "未取得"

            prod_no = None
            tag = soup.select_one("#osmPrdNo")
            if tag:
                prod_no = tag.text.strip()
            if not prod_no:
                li_tags = soup.select("li.tvlogo, li.goods-code-container")
                for tag in li_tags:
                    if "品號：" in tag.text:
                        prod_no = tag.text.split("品號：")[-1].strip()
                        break
            if not prod_no:
                meta_code = soup.select_one("meta[name='keywords']")
                if meta_code and "品號：" in meta_code.get("content", ""):
                    prod_no = meta_code["content"].split("品號：")[-1].split(",")[0].strip()
            if not prod_no:
                match = re.search(r"品號[:： ]?\s*(\w+)", soup.get_text())
                prod_no = match.group(1).strip() if match else "未取得"

            zhigui_value = _parse_bsmi_from_soup(soup)
            break

        except (requests.Timeout, requests.ConnectionError) as e:
            print(f"❌ (Mobile) 第 {attempt} 次失敗：{e}")
            if attempt < max_retries:
                time.sleep(3)
            else:
                pass
        except Exception as e:
            print(f"❌ (Mobile) 發生例外：{e}")
            break

    return {"商品名稱": name, "品號": prod_no, "商檢字號": zhigui_value, "網址": url}


# ➤ 擷取網址並處理 &amp; 解碼（保留以防後續擴充使用）
def extract_urls_from_text(text: str) -> List[str]:
    clean = html.unescape(html.unescape(text))
    urls = re.findall(r"https?://[^\s\]]+", clean)
    return list(dict.fromkeys(urls))


def to_roc_date(date_obj: datetime.date) -> str:
    roc_year = date_obj.year - 1911
    return f"{roc_year}/{date_obj.month}/{date_obj.day}"


def fetch_momo_product(url: str) -> Dict[str, str]:
    return parse_momo_simple(url, max_retries=5)


def load_input_csv() -> pd.DataFrame:
    try:
        from google.colab import files  # type: ignore

        uploaded = files.upload()
        if not uploaded:
            raise FileNotFoundError("No file uploaded.")
        filename = next(iter(uploaded))
    except Exception:
        filename = input("請輸入 CSV 檔名：").strip()
    df = pd.read_csv(filename)
    if "序號" not in df.columns or "商品網址" not in df.columns:
        raise ValueError("CSV 必須包含『序號』與『商品網址』欄位")
    return df


def build_output_rows(df: pd.DataFrame, roc_date: str) -> pd.DataFrame:
    rows: List[Dict[str, str]] = []
    for _, row in df.iterrows():
        url = str(row.get("商品網址", "")).strip()
        product_info = fetch_momo_product(url) if url else {}

        name = product_info.get("商品名稱", "") if isinstance(product_info, dict) else ""
        prod_no = product_info.get("品號", "") if isinstance(product_info, dict) else ""
        zhigui = product_info.get("商檢字號", "") if isinstance(product_info, dict) else ""
        error_msg = ""
        if not name or name.startswith("錯誤："):
            error_msg = name.replace("錯誤：", "").strip() if name else "抓取失敗"
            name = ""

        rows.append(
            {
                "編號": row.get("序號", ""),
                "檢查案號": "",
                "查核日期": roc_date,
                "網路名稱/店家名稱": "momo購物網",
                "賣家帳號或拍賣代碼": prod_no,
                "商品名稱": name,
                "再查核日期": "",
                "是否下架": "",
                "是否改正": "",
                "調查結果": error_msg,
                "網址/地址": url,
                "商檢標識": zhigui,
                "已宣導": "",
                "已下架": "",
            }
        )
        time.sleep(5)
    columns = [
        "編號",
        "檢查案號",
        "查核日期",
        "網路名稱/店家名稱",
        "賣家帳號或拍賣代碼",
        "商品名稱",
        "再查核日期",
        "是否下架",
        "是否改正",
        "調查結果",
        "網址/地址",
        "商檢標識",
        "已宣導",
        "已下架",
    ]
    return pd.DataFrame(rows, columns=columns)


def export_to_excel(df: pd.DataFrame, roc_date: str) -> str:
    filename = f"momo_check_output_ROC{roc_date.replace('/', '')}.xlsx"
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        df.to_excel(writer, index=False)
    from openpyxl import load_workbook

    wb = load_workbook(filename)
    ws = wb.active

    column_widths = [6, 15, 12, 16, 18, 30, 12, 10, 10, 30, 40, 14, 10, 10]
    for idx, width in enumerate(column_widths, start=1):
        col_letter = get_column_letter(idx)
        ws.column_dimensions[col_letter].width = width

    wrap_columns = {6, 10, 11}  # 商品名稱、調查結果、網址/地址
    for row in ws.iter_rows(min_row=1, max_col=ws.max_column, max_row=ws.max_row):
        for cell in row:
            if cell.col_idx in wrap_columns:
                cell.alignment = Alignment(wrap_text=True)

    wb.save(filename)

    try:
        from google.colab import files  # type: ignore

        files.download(filename)
    except Exception:
        print(f"已匯出檔案：{filename}")
    return filename


def main():
    df_in = load_input_csv()
    roc_date = to_roc_date(datetime.date.today())
    df_out = build_output_rows(df_in, roc_date)
    export_to_excel(df_out, roc_date)


if __name__ == "__main__":
    main()
