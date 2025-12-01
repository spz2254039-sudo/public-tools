# -*- coding: utf-8 -*-
"""
 m o m o _ c h e c k _ v 4
 在 v1/v2/v3 基礎上改為 CSV 匯入、Excel 匯出（保留 parse_momo_simple 爬蟲邏輯，去 Word 化）
"""

import datetime
import html
import re
import time
from typing import Dict, List

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

MRDT_REGEX = re.compile(r"\b[MRDT][A-Za-z0-9]{5}\b", re.IGNORECASE)


# ➤ 對 MOMO 商品進行 retry + timeout 的簡易爬蟲（保留 v1/v3 邏輯並更新商檢字號）
def parse_momo_simple(url: str, max_retries: int = 3) -> Dict[str, str]:
    headers = {"User-Agent": "Mozilla/5.0"}
    for attempt in range(1, max_retries + 1):
        try:
            res = requests.get(url, headers=headers, timeout=20)
            res.raise_for_status()
            soup = BeautifulSoup(res.text, "html.parser")

            # 商品名稱
            name_tag = soup.select_one("meta[property='og:title']")
            name = name_tag["content"].strip() if name_tag else "未取得"

            # 品號（多重嘗試）
            prod_no = None
            tag = soup.select_one("#osmPrdNo")
            if tag:
                prod_no = tag.text.strip()
            if not prod_no:
                li_tags = soup.select("li.tvlogo, li.goods-code-container")
                for tag in li_tags:
                    if "品號：" in tag.text:
                        prod_no = tag.text.split("品號：")[-1].strip()
                        break
            if not prod_no:
                meta_code = soup.select_one("meta[name='keywords']")
                if meta_code and "品號：" in meta_code["content"]:
                    prod_no = meta_code["content"].split("品號：")[-1].split(",")[0].strip()
            if not prod_no:
                match = re.search(r"品號[:： ]?\s*(\w+)", soup.get_text())
                prod_no = match.group(1).strip() if match else "未取得"

            # 商檢字號
            zhigui_value = ""

            spec_panel = soup.find("div", id="panel-2")
            if spec_panel is not None:
                label_div = spec_panel.find(
                    lambda tag: tag.name == "div" and tag.get_text(strip=True) == "商檢字號"
                )
                if label_div is not None:
                    value_div = label_div.find_next_sibling("div")
                    if value_div is not None:
                        text = value_div.get_text(" ", strip=True)
                        m = MRDT_REGEX.search(text)
                        if m:
                            zhigui_value = m.group(0).upper()

            if not zhigui_value and spec_panel is not None:
                text = spec_panel.get_text(" ", strip=True)
                m = MRDT_REGEX.search(text)
                if m:
                    zhigui_value = m.group(0).upper()

            if not zhigui_value:
                page_text = soup.get_text(" ", strip=True)
                m = MRDT_REGEX.search(page_text)
                if m:
                    zhigui_value = m.group(0).upper()

            return {"商品名稱": name, "品號": prod_no, "商檢字號": zhigui_value, "網址": url}

        except Exception as e:  # pragma: no cover - 連線異常時輸出錯誤訊息
            print(f"❌ 第 {attempt} 次失敗：{e}")
            if attempt < max_retries:
                time.sleep(3)
            else:
                return {"商品名稱": f"錯誤：{e}", "品號": "錯誤", "商檢字號": "", "網址": url}


# ➤ 擷取網址並處理 &amp; 解碼（保留以防後續擴充使用）
def extract_urls_from_text(text: str) -> List[str]:
    clean = html.unescape(html.unescape(text))
    urls = re.findall(r"https?://[^\s\]]+", clean)
    return list(dict.fromkeys(urls))


def to_roc_date(date_obj: datetime.date) -> str:
    roc_year = date_obj.year - 1911
    return f"{roc_year}/{date_obj.month}/{date_obj.day}"


def fetch_momo_product(url: str) -> Dict[str, str]:
    return parse_momo_simple(url, max_retries=3)


def load_input_csv() -> pd.DataFrame:
    try:
        from google.colab import files  # type: ignore

        uploaded = files.upload()
        if not uploaded:
            raise FileNotFoundError("No file uploaded.")
        filename = next(iter(uploaded))
    except Exception:
        filename = input("請輸入 CSV 檔名：").strip()
    df = pd.read_csv(filename)
    if "序號" not in df.columns or "商品網址" not in df.columns:
        raise ValueError("CSV 必須包含『序號』與『商品網址』欄位")
    return df


def build_output_rows(df: pd.DataFrame, roc_date: str) -> pd.DataFrame:
    rows: List[Dict[str, str]] = []
    for _, row in df.iterrows():
        url = str(row.get("商品網址", "")).strip()
        product_info = fetch_momo_product(url) if url else {}

        name = product_info.get("商品名稱", "") if isinstance(product_info, dict) else ""
        prod_no = product_info.get("品號", "") if isinstance(product_info, dict) else ""
        zhigui = product_info.get("商檢字號", "") if isinstance(product_info, dict) else ""
        error_msg = ""
        if not name or name.startswith("錯誤："):
            error_msg = name.replace("錯誤：", "").strip() if name else "抓取失敗"
            name = ""

        rows.append(
            {
                "編號": row.get("序號", ""),
                "檢查案號": "",
                "查核日期": roc_date,
                "網路名稱/店家名稱": "momo購物網",
                "賣家帳號或拍賣代碼": prod_no,
                "商品名稱": name,
                "再查核日期": "",
                "是否下架": "",
                "是否改正": "",
                "調查結果": error_msg,
                "網址/地址": url,
                "商檢標識": zhigui,
                "已宣導": "",
                "已下架": "",
            }
        )
        time.sleep(5)
    columns = [
        "編號",
        "檢查案號",
        "查核日期",
        "網路名稱/店家名稱",
        "賣家帳號或拍賣代碼",
        "商品名稱",
        "再查核日期",
        "是否下架",
        "是否改正",
        "調查結果",
        "網址/地址",
        "商檢標識",
        "已宣導",
        "已下架",
    ]
    return pd.DataFrame(rows, columns=columns)


def export_to_excel(df: pd.DataFrame, roc_date: str) -> str:
    filename = f"momo_check_output_ROC{roc_date.replace('/', '')}.xlsx"
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        df.to_excel(writer, index=False)
    from openpyxl import load_workbook

    wb = load_workbook(filename)
    ws = wb.active

    column_widths = [6, 15, 12, 16, 18, 30, 12, 10, 10, 30, 40, 14, 10, 10]
    for idx, width in enumerate(column_widths, start=1):
        col_letter = get_column_letter(idx)
        ws.column_dimensions[col_letter].width = width

    wrap_columns = {6, 10, 11}  # 商品名稱、調查結果、網址/地址
    for row in ws.iter_rows(min_row=1, max_col=ws.max_column, max_row=ws.max_row):
        for cell in row:
            if cell.col_idx in wrap_columns:
                cell.alignment = Alignment(wrap_text=True)

    wb.save(filename)

    try:
        from google.colab import files  # type: ignore

        files.download(filename)
    except Exception:
        print(f"已匯出檔案：{filename}")
    return filename


def main():
    df_in = load_input_csv()
    roc_date = to_roc_date(datetime.date.today())
    df_out = build_output_rows(df_in, roc_date)
    export_to_excel(df_out, roc_date)


if __name__ == "__main__":
    main()
